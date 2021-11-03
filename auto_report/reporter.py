from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl.styles import Font
from .constant_value import *
from PyQt5.QtCore import QObject, pyqtSignal
from copy import copy
# import ptvsd

'''
Check value of a cell contain a attr of chime and return "key" of chime dict had handled in handle_file
@params:
    cell: a cell in workbook
    type: type of report is 'Osc' or 'Logic'
@return:
    "key" of chime dict
    -1 if cell is not contain attr of chime
'''
def check_chime_attr_in_sheet(cell, type):
    if type == OSC_TYPE:
        if FREQ_TEXT in cell.value:
            return FREQ_IMG_COORD
        elif ATTACK_TEXT in cell.value:
            return ATTACK_IMG_COORD
        elif PEAK_TEXT in cell.value:
            return PEAK_IMG_COORD
        elif DECAY_TEXT in cell.value and DECAY_WITHIN_TEXT in cell.value:
            return DECAY_WITHIN_IMG_COORD
        elif FALL_TEXT in cell.value:  # or (DECAY_TEXT in cell.value and DECAY_AFTER_TEXT in cell.value):
            return FALL_IMG_COORD
        elif ZERO_GAIN_TEXT in cell.value:
            return ZERO_GAIN_IMG_COORD
        elif TOTAL_DURATION_TEXT in cell.value:
            return TOTAL_DURATION_IMG_COORD
        elif AMPLITUDE_TEXT in cell.value:
            return AMPLITUDE_IMG_COORD
        else:
            return -1
    elif type == LOGIC_TYPE:
        if FREQ_LOGIC_TEXT in cell.value:
            return FREQ_IMG_COORD
        elif ATTACK_LOGIC_TEXT in cell.value:
            return ATTACK_IMG_COORD
        elif PEAK_LOGIC_TEXT in cell.value:
            return PEAK_IMG_COORD
        elif DECAY_WITHIN_LOGIC_TEXT in cell.value:
            return DECAY_WITHIN_IMG_COORD
        elif FALL_LOGIC_TEXT in cell.value:  # or (DECAY_TEXT in cell.value and DECAY_AFTER_TEXT in cell.value):
            return FALL_IMG_COORD
        elif ZERO_LOGIC_TEXT in cell.value:
            return ZERO_GAIN_IMG_COORD
        elif TOTAL_DURATION_LOGIC_TEXT in cell.value:
            return TOTAL_DURATION_IMG_COORD
        elif AMPLITUDE_LOGIC_TEXT in cell.value:
            return AMPLITUDE_IMG_COORD
        else:
            return -1  
    elif type == WAVEFORM_TYPE:
        if CHIME_DURATION_TEXT in cell.value:
            return WAVEFORM_IMG_COORD
        else:
            return -1
    else:
        return -1

class Reporter(QObject):
    # emit the log to the Qt, display it to QListWidget
    logSignal = pyqtSignal(str)
    # emit the progressed of a chime in dict chimes
    progressed = pyqtSignal(int)
    # emit when finish the progress of add image to workbook
    finished = pyqtSignal()
    
    '''
    Init of class Reporter
    @params:
        - wb_path: path to the file workbook
        - wb_save_path: path to the file workbook shall to save
        - sheettab_color: which color will set for sheet after finish
        - is_fill_actual: shall fill actual value or not
        - is_delete_all_img: delete all images in current sheet or just delete image is update
        - report_type: 'Osc' or 'Logic'
        - dict: dict of chime from handle_files.py
    '''
    def __init__(self, wb_path, wb_save_path, sheettab_color, is_fill_actual, is_delete_all_img, report_type, dict):
        super().__init__()
        self._wb_path = wb_path
        self._wb_save_path = wb_save_path
        self._sheettab_color = sheettab_color
        self._is_fill_actual = is_fill_actual
        self._is_delete_all_img = is_delete_all_img
        self._chimes = dict
        self._report_type = report_type

    def run_report(self):
        
        # ptvsd.debug_this_thread() # uncomment when debug

        # load workbook
        wb = load_workbook(self._wb_path)
        self.logSignal.emit("Open workbook " + self._wb_path)
        self.logSignal.emit("++++++++++++++++++++++++++++++++++++++++++++++++")
        
        # copy waveform template
        # if WAVEFORM_TYPE:
        #     waveform_sheet = wb["waveform_template"]
        #     waveform_template_data = self.copy_range(start_col=1, start_row= 1, end_col=16, end_row= 32, current_sheet = waveform_sheet)

        for index, chime in enumerate(self._chimes):
            self.progressed.emit(index)
            sheet_name = '{}_{}'.format(chime[:5], chime[5:])
            if sheet_name in wb:
                sheet = wb[sheet_name]
            elif ZONE in sheet_name:
                if CHIME_16 in sheet_name and CHIME_16 in wb:
                    template_sheet = wb[CHIME_16]
                elif CHIME_17 in sheet_name and CHIME_17 in wb:
                    template_sheet = wb[CHIME_17]
                else:
                    self.logSignal.emit(' - Sheet {} not exist!!!!'.format(sheet_name))
                    continue
                # copy worksheet
                sheet = wb.copy_worksheet(template_sheet)
                # add condition formatting
                sheet.conditional_formatting = template_sheet.conditional_formatting
                sheet.title = sheet_name
            else:
                self.logSignal.emit(' - Sheet {} not exist!!!!'.format(sheet_name))
                continue
            # print("Open sheet : " + chime)
            self.logSignal.emit(" - Open sheet: " + sheet.title)

            # dict contain coord of chime attr in current sheet
            dict_coord_attr = {}
            # get total index of current chime
            total_index = len(self._chimes[chime])
            # update index in dict
            for i in range(0, total_index):
                dict_coord_attr[str(i + 1)] = {}
            # if WAVEFORM_TYPE:
            #     waveform_row = self.find_first_row_report(sheet)
            #     sheet.insert_rows(waveform_row - 1, 33)
            #     self.paste_range(start_col=1, start_row=waveform_row, end_col=16, end_row=waveform_row + 32, sheet_received=sheet, copied_data=waveform_template_data)
            #     box = (1, waveform_row, 6, waveform_row + 10)
            if self._report_type == OSC_TYPE:
                # box = (min_col, min_row, max_col, max_row)
                box = (1, 1, 101, 30 * total_index + 1)
                self.find_img_coord_in_sheet(current_sheet=sheet, dict_coord_attr=dict_coord_attr, box=box)
            elif self._report_type == WAVEFORM_TYPE:
                waveform_row = self.find_first_row_report(sheet)
                box = (1, waveform_row, 6, waveform_row + 10)
                self.find_img_coord_in_sheet(current_sheet=sheet, dict_coord_attr=dict_coord_attr, box=box)
            elif self._report_type == LOGIC_TYPE:
                self.find_logic_img_coord_in_sheet(current_sheet=sheet, dict_coord_attr=dict_coord_attr)
            # remove images
            if self._report_type != WAVEFORM_TYPE:
                if self._is_delete_all_img:
                    self.logSignal.emit("   + Remove all images in current sheet.....")
                    sheet._images = []
                else:
                    dict_old_img_coord = self.get_dict_old_img_coord(list_coord_old_img=self.find_old_img_coord(current_sheet=sheet), dict_coord_img=dict_coord_attr)
                    list_coord_shall_remove = self.find_coord_shall_remove_in_old_img(dict_old_img_coord, chime=chime)
                    for coord in list_coord_shall_remove:
                        try:
                            sheet._images.remove(coord)
                        except:
                            pass
            if self._report_type == OSC_TYPE or self._report_type == WAVEFORM_TYPE:
                self.add_images_and_set_actual_values(chime=chime, total_index=total_index, current_sheet=sheet, dict_img_coord=dict_coord_attr)
            elif self._report_type == LOGIC_TYPE:
                self.add_images_only_for_logic(chime=chime, current_sheet=sheet, dict_img_coord=dict_coord_attr)
            self.logSignal.emit("   + Set sheet color to " + self._sheettab_color)
            sheet.sheet_properties.tabColor = self._sheettab_color
        self.logSignal.emit("++++++++++++++++++++++++++++++++++++++++++++++++")
        self.logSignal.emit("Save workbook " + self._wb_save_path)
        try:
            wb.save(self._wb_save_path)
            self.logSignal.emit("Finished")
            self.finished.emit()
        except PermissionError:
            self.logSignal.emit("Error: File {} in used!!!".format(self._wb_save_path))


    def copy_range(self, start_col, start_row, end_col, end_row, current_sheet):
        range_selected = []
        for i in range(start_row, end_row + 1, 1):
            row_selected = []
            for j in range(start_col, end_col + 1, 1):
                row_selected.append(current_sheet.cell(row = i, column = j))
            range_selected.append(row_selected)
        return range_selected

    def paste_range(self, start_col, start_row, end_col, end_row, sheet_received, copied_data):
        count_row = 0
        for i in range(start_row, end_row + 1, 1):
            count_col = 0
            for j in range(start_col, end_col + 1, 1):
                sheet_received.cell(row = i, column = j).value = copied_data[count_row][count_col].value
                if copied_data[count_row][count_col].has_style:
                    sheet_received.cell(row = i, column = j).font = copy(copied_data[count_row][count_col].font)
                    sheet_received.cell(row = i, column = j).border = copy(copied_data[count_row][count_col].border)
                    sheet_received.cell(row = i, column = j).fill = copy(copied_data[count_row][count_col].fill)
                    sheet_received.cell(row = i, column = j).number_format = copy(copied_data[count_row][count_col].number_format)
                    sheet_received.cell(row = i, column = j).protection = copy(copied_data[count_row][count_col].protection)
                    sheet_received.cell(row = i, column = j).alignment = copy(copied_data[count_row][count_col].alignment)
                count_col += 1
            count_row += 1

    '''
    Get coordinate of images in current sheet
    @params:
        current_sheet
    @return:
        list of coord of images
    '''
    def find_old_img_coord(self, current_sheet):
        all_img_anchor = []
        for img in current_sheet._images:
            img_anchor = get_column_letter(img.anchor._from.col + 1) + str(img.anchor._from.row + 1)
            all_img_anchor.append(img_anchor)
        return all_img_anchor


    '''
    Compare list coord of image and dict of coord of all image
    Return a dict of "key" of attr chime and coord of image
    @params:
        list_coord_old_img: list coord of images get from method find_old_img_coord()
        dict_coord_img: dict of coord of all images get from method find_img_coord_in_sheet()
    @return:
        dict of key and value (coord) of all image in current sheet
    '''
    def get_dict_old_img_coord(self, list_coord_old_img, dict_coord_img):
        dict_old_img_coord = {}
        for current_index in dict_coord_img:
            dict_old_img_coord.update({current_index: {}})
            for img_key in dict_coord_img[current_index]:
                coord_new_img = dict_coord_img[current_index][img_key] 
                if coord_new_img in list_coord_old_img:
                    dict_old_img_coord[current_index][img_key] = coord_new_img
        return dict_old_img_coord


    '''
    
    '''
    def find_coord_shall_remove_in_old_img(self, dict_old_img, chime):
        list_coord_shall_remove = []
        for chime_index in self._chimes[chime]:
            for img_attr in self._chimes[chime][chime_index]:
                if img_attr in dict_old_img[chime_index]:
                    self.logSignal.emit("   + Remove image {} of index {}".format(img_attr, chime_index))
                    list_coord_shall_remove.append(dict_old_img[chime_index][img_attr])
        return list_coord_shall_remove

    def find_img_coord_in_sheet(self, current_sheet, dict_coord_attr, box):
        current_index = 0
        (min_col, min_row, max_col, max_row) = box
        for row in current_sheet.iter_rows(min_row=min_row,
                                    max_row=max_row,
                                    min_col=min_col,
                                    max_col=max_col):
            if row[0].row >= 30 * current_index:
                current_index += 1
            if self._report_type == WAVEFORM_TYPE:
                current_index = 1
            for cell in row:
                if cell.value == LOGIC_ANALYZER_SECTION:
                    break
                if cell.value and isinstance(cell.value, str) and PNG_EXT_TXT not in cell.value:
                    chime_attr_coord = check_chime_attr_in_sheet(cell, self._report_type)
                    if chime_attr_coord != -1:
                        # get and set coord
                        coord = '{}{}'.format(cell.column_letter, str(cell.row + 2))
                        dict_coord_attr[str(current_index)][chime_attr_coord] = coord
                    else:
                        pass
                else:
                    pass
    
    def find_logic_img_coord_in_sheet(self, current_sheet, dict_coord_attr):
        row_start = self.find_first_row_report(current_sheet=current_sheet)
        for row in current_sheet.iter_rows(min_row=row_start,
                                    max_row=row_start + 200,
                                    min_col=1,
                                    max_col=30):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and PNG_EXT_TXT not in cell.value:
                    chime_attr_coord = check_chime_attr_in_sheet(cell, self._report_type)
                    if chime_attr_coord != -1:
                        # get and set coord
                        for current_index in dict_coord_attr:
                            coord = '{}{}'.format(get_column_letter(cell.column + 20 * (int(current_index) -1)  + 2), str(cell.row + 1))
                            dict_coord_attr[str(current_index)][chime_attr_coord] = coord
                        if chime_attr_coord == AMPLITUDE_IMG_COORD:
                            break
                    else:
                        pass
                else:
                        pass

    def add_images_and_set_actual_values(self, chime, total_index, current_sheet, dict_img_coord):
        for chime_index in self._chimes[chime]:
            for img_coord in self._chimes[chime][chime_index]:
                # open chime attr image
                img = Image(self._chimes[chime][chime_index][img_coord][PATH])
                if img_coord == AMPLITUDE_IMG_COORD:
                    img.height = 304
                    img.width = 1470
                if img_coord in dict_img_coord[chime_index]:
                    # add image to worksheet
                    current_sheet.add_image(img, dict_img_coord[chime_index][img_coord])
                    # get image coord
                    image_coord = coordinate_from_string(dict_img_coord[chime_index][img_coord])
                    # set image basename coord
                    if img_coord == AMPLITUDE_IMG_COORD:
                        image_basename_coord = ''.join((image_coord[0], str(image_coord[1] + 16))) 
                    else:                  
                        image_basename_coord = ''.join((image_coord[0], str(image_coord[1] + 26)))

                    # set 'Actual:' cell value
                    if self._is_fill_actual:
                        col_img_coord = column_index_from_string(image_coord[0])
                        col_actual_coord = col_img_coord + 1
                        if img_coord == TOTAL_DURATION_IMG_COORD and total_index == 1:
                            col_actual_coord += 1
                        if img_coord == FALL_IMG_COORD:
                            col_actual_coord += 1
                        else:
                            pass
                        row_actual_coord = image_coord[1] - 1
                        actual_cell_coord = get_column_letter(col_actual_coord) + str(row_actual_coord)
                        if current_sheet[actual_cell_coord].value == ACTUAL_TEXT:
                            col_actual_value_coord = col_actual_coord + 1
                            actual_value_cell_coord = get_column_letter(col_actual_value_coord) + str(row_actual_coord)
                            if img_coord == FREQ_IMG_COORD:
                                current_sheet[actual_value_cell_coord].value = self._chimes[chime][chime_index][img_coord][FREQ]
                            else:
                                current_sheet[actual_value_cell_coord].value = self._chimes[chime][chime_index][img_coord][DURATION]
                        else:
                            pass
                    else:
                        pass
                    # add image basename to worksheet
                    current_sheet[image_basename_coord].value = self._chimes[chime][chime_index][img_coord][BASENAME]
                    self.logSignal.emit("   + Insert image " + img_coord + " of index {} at coordinate ".format(chime_index) +
                            dict_img_coord[chime_index][img_coord])
                else:
                    pass
    
    def add_images_only_for_logic(self, chime, current_sheet, dict_img_coord):
        for chime_index in self._chimes[chime]:
            for img_coord in self._chimes[chime][chime_index]:
                # open chime attr image
                img = Image(self._chimes[chime][chime_index][img_coord][PATH])
                img.height = 283
                img.width = 1127
                
                if img_coord in dict_img_coord[chime_index]:
                    # add image to worksheet
                    current_sheet.add_image(img, dict_img_coord[chime_index][img_coord])
                    # get image coord
                    image_coord = coordinate_from_string(dict_img_coord[chime_index][img_coord])
                    # set image basename coord
                    image_basename_coord = ''.join((image_coord[0], str(image_coord[1] + 14)))
                    # add image basename to worksheet
                    current_sheet[image_basename_coord].value = self._chimes[chime][chime_index][img_coord][BASENAME]
                    current_sheet[image_basename_coord].font = Font(bold=False)
                    self.logSignal.emit("   + Insert image " + img_coord + " of index {} at coordinate ".format(chime_index) +
                            dict_img_coord[chime_index][img_coord])
                else:
                    pass

    '''
    Loop through each cell of sheet and return the row of Logic analyzer section
    @params:
        current_sheet: sheet need to find
    @return:
        row value
    '''
    def find_first_row_report(self, current_sheet):
        for row in current_sheet.iter_rows(min_row=1,
                                max_row=365,
                                min_col=1,
                                max_col=3):
            for cell in row:
                if self._report_type == WAVEFORM_TYPE and cell.value == WAVEFORM_SECTION:
                    return cell.row
                elif self._report_type == LOGIC_TYPE and cell.value == LOGIC_ANALYZER_SECTION:
                    return cell.row