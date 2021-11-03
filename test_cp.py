from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl.styles import Font

from copy import copy
    
    
    
    
def copy_range(start_col, start_row, end_col, end_row, current_sheet):
    range_selected = []
    for i in range(start_row, end_row + 1, 1):
        row_selected = []
        for j in range(start_col, end_col + 1, 1):
            row_selected.append(current_sheet.cell(row = i, column = j))
        range_selected.append(row_selected)
    return range_selected

def paste_range(start_col, start_row, end_col, end_row, sheet_received, copied_data):
    count_row = 0
    for i in range(start_row, end_row + 1 , 1):
        count_col = 0
        for j in range(start_col, end_col + 1, 1):
            sheet_received.cell(row = i, column = j).value = copied_data[count_row][count_col].value
            if copied_data[count_row][count_col].has_style:
                sheet_received.cell(row = i, column = j).font = copy(copied_data[count_row][count_col].font)
                sheet_received.cell(row = i, column = j).border = copy(copied_data[count_row][count_col].border)
                sheet_received.cell(row = i, column = j).fill = copy(copied_data[count_row][count_col].fill)
                sheet_received.cell(row = i, column = j).number_format = copy(copied_data[count_row][count_col].number_format)
                sheet_received.cell(row = i, column = j).protection = copy(copied_data[count_row][count_col].protection)
                sheet_received.cell(row = i, column = j).alignment = copy(copied_data[count_row][count_col].alignment)
            count_col += 1
        count_row += 1

def find_report_type(current_sheet):
    for row in current_sheet.iter_rows(min_row=1,
                            max_row=365,
                            min_col=1,
                            max_col=3):
        for cell in row:
            if cell.value == 'Waveform':
                return
            elif cell.value == 'Logic Analyzer':
                return cell.row


def main():
    WB_PATH = "D:\\MyProjects\\auto_report\\excel\\P702_SGO_14.5V_Chime_General_Testing.xlsx"
    WB_SAVE_PATH = "D:\\MyProjects\\auto_report\\excel\\P702_SGO_14.5V_Chime_General_Testing_Updated.xlsx"
    wb = load_workbook(WB_PATH)
    waveform_sheet = wb["waveform_template"]
    waveform_template_data = copy_range(start_col=1, start_row= 1, end_col=16, end_row= 32, current_sheet = waveform_sheet)
    sheet = wb['Chime_10']
    waveform_row = find_report_type(sheet)
    # for i in range(32):
    #     sheet.insert_rows(waveform_row - 1)
    merged_cell_range = sheet.merged_cells.ranges
    # for merged_cell in merged_cell_range:
    #     merged_cell.shift(0,32)
    sheet.insert_rows(idx=waveform_row - 1, amount=32)


    paste_range(start_col=1, start_row=waveform_row, end_col=16, end_row=waveform_row + 31, sheet_received=sheet, copied_data=waveform_template_data)
    wb.save(WB_SAVE_PATH)


if __name__=='__main__':
    main()